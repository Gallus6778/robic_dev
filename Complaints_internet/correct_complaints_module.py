import openpyxl
import xlsxwriter

class Info_hlr:
    def __init__(self):
        self.info_parameter = {}
    def main(self):
        xlsx_file = "Complaints_internet/dataset_internet.xlsx"
        wb_obj = openpyxl.load_workbook(xlsx_file)
        sheet = wb_obj.active

        workbook1 = xlsxwriter.Workbook(xlsx_file)
        worksheet1 = workbook1.add_worksheet()

        # Obtenir les valeurs de la derniere ligne du dataset
        subscriber_info = {}
        subscriber_info = {
            "imsi": sheet["A" + str(sheet.max_row)].value,
            "encKey": sheet["B" + str(sheet.max_row)].value,
            "algoId": sheet["C" + str(sheet.max_row)].value,
            "kdbId": sheet["D" + str(sheet.max_row)].value,
            "acsub": sheet["E" + str(sheet.max_row)].value,
            "imsiActive": sheet["F" + str(sheet.max_row)].value,
            "accTypeGSM": sheet["G" + str(sheet.max_row)].value,
            "accTypeGERAN": sheet["H" + str(sheet.max_row)].value,
            "accTypeUTRAN": sheet["I" + str(sheet.max_row)].value,
            "odboc": sheet["J" + str(sheet.max_row)].value,
            "odbic": sheet["K" + str(sheet.max_row)].value,
            "odbr": sheet["L" + str(sheet.max_row)].value,
            "odboprc": sheet["M" + str(sheet.max_row)].value,
            "odbssm": sheet["N" + str(sheet.max_row)].value,
            "odbgprs": sheet["O" + str(sheet.max_row)].value,
            "odbsci": sheet["P" + str(sheet.max_row)].value,
            "isActiveIMSI": sheet["Q" + str(sheet.max_row)].value,
            "msisdn": sheet["R" + str(sheet.max_row)].value,
            "actIMSIGprs": sheet["S" + str(sheet.max_row)].value,
            "obGprs": sheet["T" + str(sheet.max_row)].value,
            "qosProfile": sheet["U" + str(sheet.max_row)].value,
            "refPdpContextName": sheet["V" + str(sheet.max_row)].value,
            "imeisv": sheet["W" + str(sheet.max_row)].value,
            "ldapResponse": sheet["X" + str(sheet.max_row)].value,
            "Targets": sheet["Y" + str(sheet.max_row)].value}

        list_apn = ["n-internet","n-est","n-cen","n-connect","n-ada","n-nor","n-sud","n-out","n-nwt","n-ext","n-lit","n-swt"]

        if subscriber_info["ldapResponse"] == 'None':
            # self.info_parameter["ldapResponse"] = "Unknow Subscriber"
            # return self.info_parameter["ldapResponse"]
        #
            if subscriber_info["odbgprs"] != "0" or subscriber_info["odbgprs"] == "None":
                if subscriber_info["odbgprs"] != "0" and subscriber_info["odbgprs"] != "None":
                    self.info_parameter["odbgprs"] = "Barring GPRS is True in HLR"
                elif subscriber_info["odbgprs"] == "None":
                    self.info_parameter["odbgprs"] = "Barring GPRS is not defined"
            #
            if subscriber_info["refPdpContextName"] == "None" or subscriber_info["refPdpContextName"] not in list_apn:
                if subscriber_info["refPdpContextName"] == "None":
                    self.info_parameter["refPdpContextName"] = "APN is not defined"
                elif subscriber_info["refPdpContextName"] not in list_apn:
                    if subscriber_info["refPdpContextName"] == "b-connect":
                        self.info_parameter["refPdpContextName"] = "APN=b-connect, change to n-internet"
                    elif subscriber_info["refPdpContextName"] == "n-wap":
                        self.info_parameter["refPdpContextName"] = "APN=n-wap, change to n-internet"
                    elif subscriber_info["refPdpContextName"] == "n-mms":
                        self.info_parameter["refPdpContextName"] = "APN=n-mms, change to n-internet"
                    else :
                        self.info_parameter["refPdpContextName"] = "APN is "+ subscriber_info["refPdpContextName"] + "; contact assistant"
            #
            if subscriber_info["qosProfile"] == "None" or subscriber_info['qosProfile'] not in ['QoS-R7', 'GOLD', 'SILVER', 'COPPER']:
                if subscriber_info["qosProfile"] == "None":
                    self.info_parameter["qosProfile"] = "QoS is not defined"
                elif subscriber_info['qosProfile'] not in ['QoS-R7', 'GOLD', 'SILVER', 'COPPER']:
                    self.info_parameter["qosProfile"] = "QoS " + subscriber_info['qosProfile'] + " is not good, Change to QoS-R7"
            #
            if (subscriber_info['qosProfile'] in ['QoS-R7', 'GOLD', 'SILVER', 'COPPER']) and (subscriber_info["refPdpContextName"] in list_apn) and (subscriber_info["odbgprs"] == "0"):
                self.info_parameter["result"] = "Everything is ok"
            # print(subscriber_info)

            # Ecrire l'action a mener dans le fichier dataset_internet.xlsx
            values = ''
            for keys, value in self.info_parameter.items():
                values = values + keys + ":" + value + ";"

            # worksheet1.write('Y'+ str(sheet.max_row), values)
            # workbook1.close()
            sheet.cell(row=sheet.max_row, column=sheet.max_column, value=values)
            wb_obj.save(filename=xlsx_file)
            return self.info_parameter
        elif subscriber_info["ldapResponse"] != 'None':
            self.info_parameter["ldapResponse"] = "Unknow Subscriber"

            #
            values = ''
            for keys, value in self.info_parameter.items():
                values = values + keys + ":" + value + ";"

            # worksheet1.write('Y'+ str(sheet.max_row), values)
            # workbook1.close()
            sheet.cell(row=sheet.max_row, column=sheet.max_column, value=values)
            wb_obj.save(filename=xlsx_file)
            return self.info_parameter
        else:
            pass

if __name__ == "__main__":
    subscriber_info = Info_hlr()
    info_parameter = subscriber_info.main()

    print(info_parameter)
        # ----------------- Consultation parametres HLR ----------------------